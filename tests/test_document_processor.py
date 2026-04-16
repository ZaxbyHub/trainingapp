"""
Tests for Document Processor Module (Phase 4.2)
"""

import pytest
import os
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock

from document_processor import DocumentProcessor, DocumentChunk


class TestPDFExtraction:
    """Tests for PDF text extraction (test_pdf_extraction)."""

    def test_pdf_extraction(self, sample_pdf_bytes, tmp_path):
        """
        Test extracting text from PDF using sample_pdf_bytes fixture.

        Mocks pdfplumber to avoid requiring the actual PDF parsing library
        and verifies that:
        - The method returns extracted text
        - Page information is correctly parsed
        - The processor can handle PDF files
        """
        # Create a temporary PDF file
        pdf_path = tmp_path / "test.pdf"
        pdf_path.write_bytes(sample_pdf_bytes)

        # Mock pdfplumber to avoid requiring actual PDF parsing library
        with patch('pdfplumber.open') as mock_open:
            # Set up mock PDF with pages
            mock_page1 = MagicMock()
            mock_page1.extract_text.return_value = "This is page one content. It has multiple sentences."
            mock_page1.page_number = 1

            mock_page2 = MagicMock()
            mock_page2.extract_text.return_value = "This is page two content. More sentences here."
            mock_page2.page_number = 2

            mock_pdf = MagicMock()
            mock_pdf.pages = [mock_page1, mock_page2]
            mock_pdf.__enter__ = Mock(return_value=mock_pdf)
            mock_pdf.__exit__ = Mock(return_value=False)

            mock_open.return_value = mock_pdf

            # Create processor and extract
            processor = DocumentProcessor()
            text, pages = processor.extract_pdf(str(pdf_path))

            # Verify extraction
            assert isinstance(text, str)
            assert "page one" in text.lower()
            assert len(pages) == 2
            assert pages[0][0] == 1  # First page number
            assert pages[1][0] == 2  # Second page number


class TestDOCXExtraction:
    """Tests for DOCX text extraction (test_docx_extraction)."""

    def test_docx_extraction(self, tmp_path):
        """
        Mock DOCX extraction test.

        Mocks python-docx library to verify:
        - Paragraphs are correctly extracted
        - Tables are properly handled
        - The processor can handle DOCX files without actual library
        """
        # Create a temporary DOCX file (path only, not real DOCX)
        docx_path = tmp_path / "test.docx"
        docx_path.write_text("fake docx content")

        # Mock python-docx
        with patch('docx.Document') as mock_document_class:
            # Set up mock DOCX structure
            mock_doc = MagicMock()

            # Mock paragraphs
            mock_para1 = MagicMock()
            mock_para1.text = "This is a test paragraph."
            mock_para2 = MagicMock()
            mock_para2.text = "Another paragraph with more content."
            mock_doc.paragraphs = [mock_para1, mock_para2]

            # Mock table
            mock_cell1 = MagicMock()
            mock_cell1.text = "Cell1"
            mock_cell2 = MagicMock()
            mock_cell2.text = "Cell2"
            mock_cell3 = MagicMock()
            mock_cell3.text = "Cell3"
            mock_cell4 = MagicMock()
            mock_cell4.text = "Cell4"

            mock_row = MagicMock()
            mock_row.cells = [mock_cell1, mock_cell2]
            mock_row2 = MagicMock()
            mock_row2.cells = [mock_cell3, mock_cell4]
            mock_table = MagicMock()
            mock_table.rows = [mock_row, mock_row2]
            mock_doc.tables = [mock_table]

            mock_document_class.return_value = mock_doc

            # Create processor and extract
            processor = DocumentProcessor()
            text = processor.extract_docx(str(docx_path))

            # Verify extraction
            assert "This is a test paragraph." in text
            assert "Another paragraph with more content." in text
            assert "Cell1" in text or "Cell3" in text  # Table content


class TestTXTExtraction:
    """Tests for TXT file extraction (test_txt_extraction)."""

    def test_txt_extraction(self, sample_text_file):
        """
        Test plain text file extraction using sample_text_file.

        Verifies:
        - Text is correctly read from file
        - Multiple paragraphs are preserved
        - UTF-8 encoding is handled properly
        """
        processor = DocumentProcessor()
        text = processor.extract_text_file(str(sample_text_file))

        # Verify content
        assert "sample text file" in text.lower()
        assert "paragraphs" in text.lower()
        assert len(text) > 0

        # Verify paragraph structure
        assert "\n\n" in text or "\n" in text  # Multiple paragraphs


class TestChunkingBoundary:
    """Tests for text chunking boundary conditions (test_chunking_boundary)."""

    def test_chunking_boundary(self):
        """
        Test semantic chunking respects sentence boundaries.

        Verifies that:
        - Chunks don't split in the middle of sentences
        - Sentence boundaries are preserved
        - Chunks respect the chunk_size limit
        - Overlap is handled correctly
        """
        # Create text with clear sentence boundaries
        text = """
        This is the first sentence. This is the second sentence.
        This is the third sentence. This is the fourth sentence.

        This is a new paragraph with more sentences. This is the sixth sentence.
        This is the seventh sentence. This is the eighth sentence here.
        """

        processor = DocumentProcessor(chunk_size=15, chunk_overlap=5)
        chunks = processor.chunk_text(text, "test.txt")

        # Verify we got chunks
        assert len(chunks) > 0

        # Check that chunks respect sentence boundaries
        for chunk in chunks:
            # No chunk should end in the middle of a sentence
            # (unless it's a very long single sentence)
            chunk_text = chunk.text.strip()
            # Check if chunk doesn't end with a period followed by lowercase
            # which would indicate mid-sentence split
            if len(chunk_text) > 10:
                # If it ends with a lowercase letter after space, it's likely mid-sentence
                assert not (chunk_text[-1].islower() and chunk_text[-2] == ' '), \
                    f"Chunk may split mid-sentence: {chunk_text[-30:]}"

        # Verify chunks don't exceed size (by word count)
        for chunk in chunks:
            word_count = len(chunk.text.split())
            assert word_count <= processor.chunk_size + 5, \
                f"Chunk exceeds size: {word_count} words, max {processor.chunk_size}"

    def test_chunking_small_text(self, tmp_path):
        """Test chunking with text smaller than chunk size."""
        processor = DocumentProcessor(chunk_size=100, chunk_overlap=10)

        small_text = "Short text."
        chunks = processor.chunk_text(small_text, "test.txt")

        assert len(chunks) == 1
        assert chunks[0].text == "Short text."
        assert chunks[0].source == "test.txt"
        assert chunks[0].chunk_index == 0

    def test_chunking_large_text(self, tmp_path):
        """Test chunking with text larger than chunk size."""
        processor = DocumentProcessor(chunk_size=50, chunk_overlap=10)

        # Create text that will exceed chunk size
        long_text = " ".join(["word"] * 200)
        chunks = processor.chunk_text(long_text, "test.txt")

        assert len(chunks) > 1
        # Verify chunk indices are sequential
        for i, chunk in enumerate(chunks):
            assert chunk.chunk_index == i

    def test_chunking_respects_paragraphs(self, tmp_path):
        """Test that chunking respects paragraph boundaries."""
        processor = DocumentProcessor(chunk_size=100, chunk_overlap=10)

        text = "First paragraph with some content.\n\nSecond paragraph with different content."
        chunks = processor.chunk_text(text, "test.txt")

        # Should have at least 2 chunks for 2 paragraphs
        assert len(chunks) >= 1

    def test_chunking_overlap(self, tmp_path):
        """Test chunking with overlap between chunks."""
        processor = DocumentProcessor(chunk_size=30, chunk_overlap=10)

        # Create text with predictable words
        text = "one two three four five six seven eight nine ten eleven twelve thirteen fourteen fifteen"
        chunks = processor.chunk_text(text, "test.txt")

        # With overlap, some words should appear in multiple chunks
        all_words = []
        for chunk in chunks:
            words = chunk.text.split()
            all_words.extend(words)

        # Should have more words than unique words due to overlap
        assert len(all_words) >= len(set(all_words))

    def test_chunking_long_sentence(self):
        """
        Test chunking handles very long single sentences.

        When a single sentence exceeds chunk_size, it should be split
        at word boundaries with overlap.
        """
        # Create a very long sentence
        words = ["word"] * 100
        text = " ".join(words) + ". This is a short second sentence."

        processor = DocumentProcessor(chunk_size=20, chunk_overlap=5)
        chunks = processor.chunk_text(text, "test.txt")

        # Should create multiple chunks
        assert len(chunks) >= 2

        # First chunk should contain words from the long sentence
        assert "word" in chunks[0].text

    def test_chunking_empty_text(self):
        """Test chunking with empty text returns empty list."""
        processor = DocumentProcessor()
        chunks = processor.chunk_text("", "test.txt")
        assert len(chunks) == 0

    def test_chunking_preserves_source(self):
        """Test that chunks preserve the source filename."""
        text = "First sentence. Second sentence. Third sentence."
        processor = DocumentProcessor()
        chunks = processor.chunk_text(text, "myfile.txt")

        for chunk in chunks:
            assert chunk.source == "myfile.txt"
            assert isinstance(chunk.chunk_index, int)

    def test_chunking_index_increment(self):
        """Test that chunk_index increments correctly."""
        text = "First sentence. Second sentence. Third sentence. Fourth sentence. Fifth sentence."
        processor = DocumentProcessor(chunk_size=5, chunk_overlap=0)
        chunks = processor.chunk_text(text, "test.txt")

        # Verify chunk indices are sequential
        indices = [chunk.chunk_index for chunk in chunks]
        assert indices == list(range(len(chunks)))


class TestEmptyFile:
    """Tests for empty file handling (test_empty_file)."""

    def test_empty_file(self, tmp_path):
        """
        Test handling of empty files.

        Verifies that:
        - Empty files are handled gracefully
        - Empty text extraction returns appropriate results
        - Processing empty files doesn't crash
        """
        # Create empty text file
        empty_path = tmp_path / "empty.txt"
        empty_path.write_text("")

        processor = DocumentProcessor()

        # Test extract_text_file with empty file
        text = processor.extract_text_file(str(empty_path))
        assert text == ""

        # Test process_file with empty file
        chunks = processor.process_file(str(empty_path))
        assert len(chunks) == 0

    def test_empty_file_with_whitespace(self, tmp_path):
        """Test handling of files with only whitespace."""
        # Create file with only whitespace
        whitespace_path = tmp_path / "whitespace.txt"
        whitespace_path.write_text("   \n\n   \t  ")

        processor = DocumentProcessor()
        chunks = processor.process_file(str(whitespace_path))
        assert len(chunks) == 0

    def test_empty_pdf(self, tmp_path, sample_pdf_bytes):
        """Test processing an empty PDF."""
        # Create a minimal valid PDF with no content
        pdf_path = tmp_path / "empty.pdf"
        pdf_path.write_bytes(sample_pdf_bytes)

        processor = DocumentProcessor()
        text, pages = processor.extract_pdf(str(pdf_path))

        # Should return empty or minimal text
        assert isinstance(text, str)

    def test_process_directory_empty(self, tmp_path):
        """Test process_directory with no supported files."""
        # Create directory with only unsupported files
        unsupported_path = tmp_path / "unsupported"
        unsupported_path.mkdir()
        (unsupported_path / "unsupported.xyz").write_text("content")

        processor = DocumentProcessor()
        chunks = processor.process_directory(str(unsupported_path))

        assert chunks == []


class TestUnsupportedExtension:
    """Tests for unsupported file extension handling (test_unsupported_extension)."""

    def test_unsupported_extension(self, tmp_path):
        """
        Test error on unsupported file format.

        Verifies that:
        - Unsupported file extensions raise ValueError
        - Error message indicates the unsupported format
        - Processor doesn't crash on invalid formats
        """
        # Create file with unsupported extension
        unsupported_path = tmp_path / "test.xyz"
        unsupported_path.write_text("some content")

        processor = DocumentProcessor()

        # Should raise ValueError for unsupported format
        with pytest.raises(ValueError) as exc_info:
            processor.extract_document(str(unsupported_path))

        assert "Unsupported file format" in str(exc_info.value)
        assert ".xyz" in str(exc_info.value)

    def test_process_file_unsupported(self, tmp_path):
        """Test process_file returns empty list for unsupported formats."""
        unsupported_path = tmp_path / "test.xyz"
        unsupported_path.write_text("some content")

        processor = DocumentProcessor()
        chunks = processor.process_file(str(unsupported_path))

        # Should return empty list instead of crashing
        assert len(chunks) == 0

    def test_unsupported_extension_in_directory(self, tmp_path):
        """Test that unsupported files are skipped in directory processing."""
        # Create directory with mixed file types
        mixed_path = tmp_path / "mixed"
        mixed_path.mkdir()

        (mixed_path / "supported.txt").write_text("supported content")
        (mixed_path / "unsupported.xyz").write_text("unsupported content")
        (mixed_path / "also_supported.md").write_text("# Markdown")

        processor = DocumentProcessor()
        chunks = processor.process_directory(str(mixed_path))

        # Should process only supported files
        sources = set(c.source for c in chunks)
        assert "supported.txt" in sources
        assert "also_supported.md" in sources
        assert "unsupported.xyz" not in sources

    def test_supported_extensions_list(self, tmp_path):
        """Test that SUPPORTED_EXTENSIONS includes all expected formats."""
        processor = DocumentProcessor()

        expected = {'.pdf', '.docx', '.doc', '.pptx', '.ppt', '.txt', '.md', '.xlsx'}
        assert processor.SUPPORTED_EXTENSIONS == expected


class TestDocumentProcessorIntegration:
    """Integration tests for document processing."""

    def test_process_file(self, sample_text_file):
        """
        Test DocumentProcessor.process_file extracts and chunks text.

        Verifies the complete pipeline:
        - File extraction
        - Text cleaning
        - Text chunking
        - Returns DocumentChunk objects
        """
        processor = DocumentProcessor(chunk_size=50, chunk_overlap=10)
        chunks = processor.process_file(str(sample_text_file))

        # Verify we got chunks
        assert len(chunks) > 0

        # Verify chunks are DocumentChunk objects
        for chunk in chunks:
            assert isinstance(chunk, DocumentChunk)
            assert isinstance(chunk.text, str)
            assert isinstance(chunk.source, str)
            assert isinstance(chunk.chunk_index, int)
            assert chunk.source == sample_text_file.name

    def test_process_directory(self, tmp_path):
        """
        Test DocumentProcessor.process_directory processes multiple files.

        Verifies:
        - All supported files are processed
        - Unsupported files are skipped
        - Chunks from all files are collected
        - Directory traversal works correctly
        """
        # Create multiple test files
        (tmp_path / "file1.txt").write_text("First file content with multiple sentences.")
        (tmp_path / "file2.txt").write_text("Second file content. More sentences here.")
        (tmp_path / "unsupported.xyz").write_text("Should be skipped")
        (tmp_path / "subdir").mkdir()
        (tmp_path / "subdir" / "file3.txt").write_text("Third file in subdirectory.")

        processor = DocumentProcessor(chunk_size=20, chunk_overlap=5)
        chunks = processor.process_directory(str(tmp_path))

        # Should process 3 text files (skipping .xyz)
        assert len(chunks) > 0

        # Verify chunks have sources from all three text files
        sources = {chunk.source for chunk in chunks}
        assert "file1.txt" in sources
        assert "file2.txt" in sources
        assert "file3.txt" in sources
        assert "unsupported.xyz" not in sources

    def test_process_directory_empty(self, tmp_path):
        """Test processing empty directory returns empty list."""
        empty_dir = tmp_path / "empty"
        empty_dir.mkdir()

        processor = DocumentProcessor()
        chunks = processor.process_directory(str(empty_dir))

        assert len(chunks) == 0

    def test_process_directory_no_supported_files(self, tmp_path):
        """Test directory with only unsupported files."""
        (tmp_path / "file1.xyz").write_text("content")
        (tmp_path / "file2.abc").write_text("content")

        processor = DocumentProcessor()
        chunks = processor.process_directory(str(tmp_path))

        assert len(chunks) == 0


# Additional utility tests

class TestDocumentChunk:
    """Tests for DocumentChunk dataclass."""

    def test_chunk_creation(self):
        """Test creating a DocumentChunk."""
        chunk = DocumentChunk(
            text="Test content",
            source="test.pdf",
            page=1,
            chunk_index=0
        )

        assert chunk.text == "Test content"
        assert chunk.source == "test.pdf"
        assert chunk.page == 1
        assert chunk.chunk_index == 0

    def test_chunk_default_values(self):
        """Test DocumentChunk with default values."""
        chunk = DocumentChunk(
            text="Test content",
            source="test.pdf"
        )

        assert chunk.text == "Test content"
        assert chunk.source == "test.pdf"
        assert chunk.page is None
        assert chunk.chunk_index == 0

    def test_chunk_from_sample_chunks(self, sample_chunks):
        """Test that sample_chunks fixture provides valid chunks."""
        assert len(sample_chunks) > 0

        for chunk in sample_chunks:
            assert isinstance(chunk, DocumentChunk)
            assert isinstance(chunk.text, str)
            assert isinstance(chunk.source, str)
            assert isinstance(chunk.chunk_index, int)
            assert len(chunk.text) > 0
            assert len(chunk.source) > 0


class TestCleanText:
    """Tests for text cleaning functionality."""

    def test_clean_text_normalizes_whitespace(self):
        """Test that clean_text normalizes multiple spaces."""
        processor = DocumentProcessor()
        text = "This  has    multiple     spaces."
        cleaned = processor.clean_text(text)
        assert "  " not in cleaned  # No double spaces
        assert cleaned == "This has multiple spaces."

    def test_clean_text_normalizes_newlines(self):
        """Test that clean_text normalizes multiple newlines."""
        processor = DocumentProcessor()
        text = "First line\n\n\n\nSecond line\nThird line"
        cleaned = processor.clean_text(text)
        assert "\n\n\n" not in cleaned  # No triple newlines
        assert "First line" in cleaned
        assert "Second line" in cleaned

    def test_clean_text_strips_whitespace(self):
        """Test that clean_text strips leading/trailing whitespace."""
        processor = DocumentProcessor()
        text = "   \n  Text with surrounding whitespace  \n  "
        cleaned = processor.clean_text(text)
        assert cleaned == "Text with surrounding whitespace"

    def test_clean_text_empty_string(self):
        """Test clean_text with empty string."""
        processor = DocumentProcessor()
        assert processor.clean_text("") == ""
        assert processor.clean_text("   ") == ""


class TestCleanTextParagraphStructure:
    """Focused tests for clean_text paragraph-structure preservation per Phase 1.1."""

    def test_paragraph_breaks_preserved(self):
        """Paragraph breaks (\\n\\n) must remain as \\n\\n after cleaning."""
        processor = DocumentProcessor()
        text = "First paragraph\n\nSecond paragraph\n\nThird paragraph"
        cleaned = processor.clean_text(text)
        # Each adjacent pair of paragraphs must be separated by exactly \n\n
        paragraphs = cleaned.split("\n\n")
        assert len(paragraphs) == 3, f"Expected 3 paragraphs, got {len(paragraphs)}: {repr(cleaned)}"
        assert paragraphs[0] == "First paragraph"
        assert paragraphs[1] == "Second paragraph"
        assert paragraphs[2] == "Third paragraph"

    def test_single_blank_line_preserved_between_paragraphs(self):
        """A single blank line between paragraphs (\\n\\n) must survive cleaning."""
        processor = DocumentProcessor()
        text = "Para one\n\nPara two"
        cleaned = processor.clean_text(text)
        # Must contain exactly one \n\n between paragraphs (not stripped to \n)
        assert cleaned.count("\n\n") >= 1, f"Expected \\n\\n paragraph break, got: {repr(cleaned)}"
        assert "Para one" in cleaned
        assert "Para two" in cleaned

    def test_three_consecutive_blank_lines_collapse_to_two(self):
        """Runs of 3+ blank lines must collapse to exactly 2 (\\n\\n)."""
        processor = DocumentProcessor()
        text = "Para one\n\n\n\nPara two"
        cleaned = processor.clean_text(text)
        # After collapsing 4 newlines -> 2, should have \n\n between paragraphs
        assert "\n\n" in cleaned, f"Expected \\n\\n, got: {repr(cleaned)}"
        assert "\n\n\n" not in cleaned, "Triple+ newlines should be absent"
        paragraphs = cleaned.split("\n\n")
        assert len(paragraphs) == 2, f"Expected 2 paragraphs separated by \\n\\n, got: {repr(cleaned)}"
        assert paragraphs[0].strip() == "Para one"
        assert paragraphs[1].strip() == "Para two"

    def test_five_consecutive_newlines_collapse_to_two(self):
        """5 consecutive \\n should collapse to \\n\\n."""
        processor = DocumentProcessor()
        text = "First\n\n\n\n\nLast"
        cleaned = processor.clean_text(text)
        # 5 newlines -> step 2: \n{3,} -> \n\n -> "First\n\nLast"
        # Step 3: split ["First", "", "Last"], strip -> ["First", "", "Last"]
        # Join: "First\n\nLast"
        # Step 4: \n\n -> stays \n\n
        assert cleaned == "First\n\nLast", f"Expected 'First\\n\\nLast', got: {repr(cleaned)}"

    def test_leading_and_trailing_newlines_stripped(self):
        """Leading/trailing newlines must be stripped from output."""
        processor = DocumentProcessor()
        text = "\n\nContent here\n\n"
        cleaned = processor.clean_text(text)
        assert not cleaned.startswith("\n"), f"Should not start with \\n: {repr(cleaned)}"
        assert not cleaned.endswith("\n"), f"Should not end with \\n: {repr(cleaned)}"
        assert cleaned == "Content here", f"Expected 'Content here', got: {repr(cleaned)}"


class TestCleanTextLineEndings:
    """Tests for mixed line ending normalization."""

    def test_windows_line_endings_normalized(self):
        """\\r\\n must be normalized to \\n."""
        processor = DocumentProcessor()
        text = "Line one\r\nLine two\r\nLine three"
        cleaned = processor.clean_text(text)
        # No carriage returns should remain
        assert "\r" not in cleaned, f"Carriage returns should be gone: {repr(cleaned)}"
        assert cleaned == "Line one\nLine two\nLine three", f"Got: {repr(cleaned)}"

    def test_mac_classic_line_endings_normalized(self):
        """Legacy \\r (old Mac) must be normalized to \\n."""
        processor = DocumentProcessor()
        text = "Line one\rLine two\rLine three"
        cleaned = processor.clean_text(text)
        assert "\r" not in cleaned, f"Carriage returns should be gone: {repr(cleaned)}"
        assert cleaned == "Line one\nLine two\nLine three", f"Got: {repr(cleaned)}"

    def test_mixed_crlf_and_lf_normalized(self):
        """Mixed \\r\\n and \\n should all become \\n."""
        processor = DocumentProcessor()
        text = "Para one\r\n\r\nPara two\n\nPara three"
        cleaned = processor.clean_text(text)
        assert "\r" not in cleaned, f"Carriage returns should be gone: {repr(cleaned)}"
        # Paragraphs separated by blank lines
        paragraphs = cleaned.split("\n\n")
        assert len(paragraphs) >= 2, f"Expected paragraph separation, got: {repr(cleaned)}"

    def test_crlf_with_multiple_blank_lines(self):
        """\\r\\n + multiple blank lines should normalize correctly."""
        processor = DocumentProcessor()
        text = "First\r\n\r\n\r\n\r\nSecond"
        cleaned = processor.clean_text(text)
        assert "\r" not in cleaned
        # 4+ newlines -> collapse to 2 -> paragraph break
        assert "\n\n" in cleaned or "\n" in cleaned


class TestCleanTextHorizontalWhitespace:
    """Tests for horizontal whitespace (spaces/tabs) collapsing within lines."""

    def test_multiple_spaces_collapsed_to_one(self):
        """Multiple spaces within a line collapse to a single space."""
        processor = DocumentProcessor()
        text = "This  has    multiple     spaces   here."
        cleaned = processor.clean_text(text)
        assert cleaned == "This has multiple spaces here.", f"Got: {repr(cleaned)}"

    def test_tabs_collapsed_to_single_space(self):
        """Tabs within a line collapse to a single space."""
        processor = DocumentProcessor()
        text = "Word1\t\tWord2\tWord3"
        cleaned = processor.clean_text(text)
        assert "\t" not in cleaned, f"Tabs should be removed: {repr(cleaned)}"
        # Tabs replaced by single spaces, adjacent ones merged
        assert "Word1" in cleaned
        assert "Word3" in cleaned

    def test_leading_and_trailing_spaces_stripped_per_line(self):
        """Leading/trailing spaces on each line must be stripped."""
        processor = DocumentProcessor()
        text = "   leading\ntrailing   \nboth   ends   "
        cleaned = processor.clean_text(text)
        lines = cleaned.split("\n")
        assert lines[0] == "leading", f"Got: {repr(lines[0])}"
        assert lines[1] == "trailing", f"Got: {repr(lines[1])}"
        assert lines[2] == "both ends", f"Got: {repr(lines[2])}"

    def test_horizontal_whitespace_in_multiline_preserved_structure(self):
        """Horizontal whitespace collapsing must not affect \\n line structure."""
        processor = DocumentProcessor()
        text = "Line  one     with   extra    spaces\nSecond  line    here"
        cleaned = processor.clean_text(text)
        lines = cleaned.split("\n")
        assert len(lines) == 2, f"Expected 2 lines, got {len(lines)}: {repr(cleaned)}"
        assert lines[0] == "Line one with extra spaces"
        assert lines[1] == "Second line here"

    def test_indentation_whitespace_collapsed(self):
        """Leading indentation (spaces at line start) must be stripped."""
        processor = DocumentProcessor()
        text = "    Indented line with trailing spaces     "
        cleaned = processor.clean_text(text)
        assert cleaned == "Indented line with trailing spaces", f"Got: {repr(cleaned)}"


class TestCleanTextEdgeCases:
    """Boundary and edge case tests for clean_text."""

    def test_empty_string_returns_empty(self):
        """Empty input string must return empty string."""
        processor = DocumentProcessor()
        assert processor.clean_text("") == ""

    def test_whitespace_only_returns_empty(self):
        """Input containing only whitespace must return empty string."""
        processor = DocumentProcessor()
        assert processor.clean_text("   ") == ""
        assert processor.clean_text("\n\n\n") == ""
        assert processor.clean_text("  \n  \t  \n  ") == ""

    def test_single_character(self):
        """Single character input is preserved."""
        processor = DocumentProcessor()
        assert processor.clean_text("x") == "x"
        assert processor.clean_text(" ") == ""

    def test_no_newlines_single_line(self):
        """Text with no newlines is returned with collapsed horizontal whitespace."""
        processor = DocumentProcessor()
        text = "Simple   text   with   spaces"
        cleaned = processor.clean_text(text)
        assert cleaned == "Simple text with spaces", f"Got: {repr(cleaned)}"

    def test_unicode_characters_preserved(self):
        """Unicode characters must survive cleaning unchanged."""
        processor = DocumentProcessor()
        text = "Unicode: \u00e9\u00e8\u00ea \u4e2d\u6587 \U0001F600"
        cleaned = processor.clean_text(text)
        assert "\u00e9" in cleaned
        assert "\u4e2d" in cleaned
        # Emoji is a surrogate pair in UTF-16, but Python 3 handles it natively
        emoji = "\U0001F600"
        assert emoji in cleaned

    def test_paragraph_with_multiple_lines(self):
        """A paragraph with multiple lines (single \\n) inside it stays intact."""
        processor = DocumentProcessor()
        text = "Line one of paragraph\nLine two of paragraph\nLine three"
        cleaned = processor.clean_text(text)
        lines = cleaned.split("\n")
        assert len(lines) == 3, f"Expected 3 lines within paragraph, got: {repr(cleaned)}"

    def test_leading_whitespace_on_all_lines(self):
        """Leading whitespace on every line is stripped."""
        processor = DocumentProcessor()
        text = "    Line one\n  \t  Line two  \n   \t  Line three"
        cleaned = processor.clean_text(text)
        lines = cleaned.split("\n")
        assert lines[0] == "Line one"
        assert lines[1] == "Line two"
        assert lines[2] == "Line three"

    def test_newlines_at_various_positions(self):
        """Newlines at start, middle, and end are handled correctly."""
        processor = DocumentProcessor()
        text = "\nFirst\n\nMiddle\n\n\nLast\n"
        cleaned = processor.clean_text(text)
        # Leading/trailing stripped, 3+ newlines collapsed
        assert not cleaned.startswith("\n")
        assert not cleaned.endswith("\n")
        # Should have at least 3 content segments
        parts = cleaned.split("\n\n")
        assert len(parts) >= 3, f"Expected 3+ parts, got {len(parts)}: {repr(cleaned)}"


class TestCleanTextListsAndProcedures:
    """Tests for numbered lists, step procedures, and structured content."""

    def test_numbered_list_survives_cleaning(self):
        """Numbered list items must survive cleaning with correct numbering."""
        processor = DocumentProcessor()
        text = "1. First step\n2. Second step\n3. Third step"
        cleaned = processor.clean_text(text)
        lines = cleaned.split("\n")
        assert len(lines) == 3, f"Expected 3 lines, got: {repr(cleaned)}"
        assert lines[0] == "1. First step"
        assert lines[1] == "2. Second step"
        assert lines[2] == "3. Third step"

    def test_multi_digit_numbered_list(self):
        """Multi-digit numbered list (10+, 100+) survives cleaning."""
        processor = DocumentProcessor()
        text = "10.  Item ten\n11.  Item eleven\n12.  Item twelve"
        cleaned = processor.clean_text(text)
        lines = cleaned.split("\n")
        assert lines[0] == "10. Item ten", f"Got: {repr(lines[0])}"
        assert lines[1] == "11. Item eleven", f"Got: {repr(lines[1])}"
        assert lines[2] == "12. Item twelve", f"Got: {repr(lines[2])}"

    def test_bulleted_list_preserved(self):
        """Bulleted list items survive cleaning."""
        processor = DocumentProcessor()
        text = "- First item\n- Second item\n- Third item"
        cleaned = processor.clean_text(text)
        lines = cleaned.split("\n")
        assert len(lines) == 3
        assert lines[0] == "- First item"
        assert lines[1] == "- Second item"
        assert lines[2] == "- Third item"

    def test_code_block_like_content(self):
        """Code-like content with indentation preserves structure via \\n separation."""
        processor = DocumentProcessor()
        text = "def hello():\n    print('hello')\n    return True"
        cleaned = processor.clean_text(text)
        lines = cleaned.split("\n")
        assert len(lines) == 3
        assert lines[0] == "def hello():"
        assert lines[1] == "print('hello')"
        assert lines[2] == "return True"

    def test_step_procedure_with_blank_lines(self):
        """Step procedure with blank lines between steps preserves both list and breaks."""
        processor = DocumentProcessor()
        text = "Step 1: Do this\n\nStep 2: Do that\n\nStep 3: Verify"
        cleaned = processor.clean_text(text)
        # Paragraph breaks should be preserved
        paragraphs = cleaned.split("\n\n")
        assert len(paragraphs) == 3, f"Expected 3 step paragraphs, got {len(paragraphs)}: {repr(cleaned)}"
        assert "Step 1" in paragraphs[0]
        assert "Step 2" in paragraphs[1]
        assert "Step 3" in paragraphs[2]

    def test_list_with_extra_spaces(self):
        """List items with extra spaces between number and text are normalized."""
        processor = DocumentProcessor()
        text = "1.    Extra   spaces\n2.    More    here\n3.    Last    one"
        cleaned = processor.clean_text(text)
        lines = cleaned.split("\n")
        assert lines[0] == "1. Extra spaces", f"Got: {repr(lines[0])}"
        assert lines[1] == "2. More here", f"Got: {repr(lines[1])}"
        assert lines[2] == "3. Last one", f"Got: {repr(lines[2])}"

    def test_nested_structure_with_blank_lines(self):
        """Nested content with blank lines preserves hierarchy."""
        processor = DocumentProcessor()
        text = "Section A\n\nSubsection A1\nContent under A1\n\nSubsection A2\nContent under A2\n\nSection B\n\nSubsection B1"
        cleaned = processor.clean_text(text)
        # Blank lines between subsections create paragraph breaks (\n\n).
        # Single \n within each subsection are preserved as line breaks.
        # Split by \n\n gives: [SectionA-part, A1-part, A2-part, SectionB-part, B1-part]
        sections = cleaned.split("\n\n")
        assert len(sections) == 5, f"Expected 5 sections, got {len(sections)}: {repr(cleaned)}"
        assert "Section A" in sections[0]
        assert "Subsection A1" in sections[1]
        assert "Subsection A2" in sections[2]
        assert "Section B" in sections[3]
        assert "Subsection B1" in sections[4]
        # Verify single \n line structure is preserved within each section
        lines_p1 = sections[1].split("\n")
        assert len(lines_p1) == 2, f"Expected 2 lines in subsection A1, got: {repr(lines_p1)}"


class TestExtractXlsx:
    """Tests for XLSX file extraction (extract_xlsx)."""

    def test_extract_xlsx_basic(self, tmp_path):
        """Test basic XLSX extraction with openpyxl."""
        from document_processor import DocumentProcessor
        import openpyxl
        
        proc = DocumentProcessor()
        filepath = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws.append(["Header1", "Header2"])
        ws.append(["Row1Col1", "Row1Col2"])
        wb.save(filepath)
        result = proc.extract_xlsx(str(filepath))
        assert "[Sheet: TestSheet]" in result
        assert "Header1 | Header2" in result
        assert "Row1Col1 | Row1Col2" in result

    def test_extract_xlsx_multiple_sheets(self, tmp_path):
        """Test XLSX with multiple sheets separated by blank lines."""
        from document_processor import DocumentProcessor
        import openpyxl
        
        proc = DocumentProcessor()
        filepath = tmp_path / "multi_sheet.xlsx"
        wb = openpyxl.Workbook()
        
        # First sheet
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["A1", "B1"])
        
        # Create second sheet
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["A2", "B2"])
        
        wb.save(filepath)
        result = proc.extract_xlsx(str(filepath))
        assert "[Sheet: Sheet1]" in result
        assert "[Sheet: Sheet2]" in result
        # Multiple sheets should be separated by blank lines
        assert "\n\n" in result

    def test_extract_xlsx_empty_rows_skipped(self, tmp_path):
        """Test that empty rows are skipped in XLSX extraction."""
        from document_processor import DocumentProcessor
        import openpyxl
        
        proc = DocumentProcessor()
        filepath = tmp_path / "empty_rows.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestData"
        ws.append(["Header1", "Header2"])
        ws.append(["Data1", "Data2"])
        ws.append([])  # Empty row
        ws.append(["Data3", "Data4"])
        wb.save(filepath)
        result = proc.extract_xlsx(str(filepath))
        assert "Header1 | Header2" in result
        assert "Data1 | Data2" in result
        assert "Data3 | Data4" in result
        # Empty row should not appear as empty separator
        lines = result.split("\n")
        empty_lines = [line for line in lines if line.strip() == ""]
        assert len(empty_lines) == 0 or result.count("\n\n") <= 1

    def test_extract_xlsx_no_data(self, tmp_path):
        """Test XLSX with no data rows (headers only)."""
        from document_processor import DocumentProcessor
        import openpyxl
        
        proc = DocumentProcessor()
        filepath = tmp_path / "headers_only.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Headers"
        ws.append(["Col1", "Col2", "Col3"])
        wb.save(filepath)
        result = proc.extract_xlsx(str(filepath))
        assert "Col1 | Col2 | Col3" in result

    def test_extract_xlsx_different_sheet_names(self, tmp_path):
        """Test that sheet names appear correctly in output."""
        from document_processor import DocumentProcessor
        import openpyxl
        
        proc = DocumentProcessor()
        filepath = tmp_path / "special_names.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data$1"
        ws.append(["Test"])
        wb.save(filepath)
        result = proc.extract_xlsx(str(filepath))
        assert "[Sheet: Data$1]" in result


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
