"""
Adversarial tests for extract_xlsx() in document_processor.py.
Tests: malformed inputs, oversized payloads, boundary violations, injection attempts.
"""

import os
import sys
import tempfile
import zipfile
import io
import pytest
import logging

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from document_processor import DocumentProcessor

logging.disable(logging.CRITICAL)


class TestExtractXlsxAdversarial:
    """Adversarial test suite for extract_xlsx()."""

    def setup_method(self):
        self.processor = DocumentProcessor(chunk_size=256, chunk_overlap=50)

    # ─── BOUNDARY: Empty / None / Invalid Type ───────────────────────────────

    def test_none_path(self):
        """None path should raise TypeError, not AttributeError."""
        with pytest.raises(TypeError):
            self.processor.extract_xlsx(None)

    def test_empty_string_path(self):
        """Empty string path should raise InvalidFileException (openpyxl rejects it)."""
        from openpyxl.utils.exceptions import InvalidFileException
        with pytest.raises(InvalidFileException):
            self.processor.extract_xlsx("")

    def test_whitespace_only_path(self):
        """Whitespace-only string should raise InvalidFileException (openpyxl rejects it)."""
        from openpyxl.utils.exceptions import InvalidFileException
        with pytest.raises(InvalidFileException):
            self.processor.extract_xlsx("   ")

    # ─── BOUNDARY: Non-existent / Missing File ────────────────────────────────

    def test_nonexistent_path(self):
        """Non-existent file should raise FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            self.processor.extract_xlsx("C:/nonexistent/fake/file_12345.xlsx")

    def test_nonexistent_relative_path(self):
        """Non-existent relative path should raise FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            self.processor.extract_xlsx("does_not_exist_987654.xlsx")

    # ─── BOUNDARY: Path Traversal / Injection ────────────────────────────────

    def test_path_traversal_dotdot(self):
        """Path traversal with .. should raise FileNotFoundError (or not escape sandbox)."""
        # Attempt to traverse upward — should be treated as a real path
        with pytest.raises(FileNotFoundError):
            self.processor.extract_xlsx("../etc/passwd.xlsx")

    def test_path_traversal_absolute_etc(self):
        """Absolute path traversal to /etc should raise FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            self.processor.extract_xlsx("/etc/passwd.xlsx")

    def test_path_traversal_windows_absolute(self):
        """Windows absolute path traversal should raise FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            self.processor.extract_xlsx("C:/Windows/System32/config.xlsx")

    def test_null_byte_injection(self):
        """Null byte in path should raise either ValueError or FileNotFoundError."""
        # Null byte (\x00) in path — openpyxl should reject it
        with pytest.raises((ValueError, FileNotFoundError)):
            self.processor.extract_xlsx("C:\\Windows\\null\x00byte.xlsx")

    # ─── BOUNDARY: Wrong file type passed as .xlsx ───────────────────────────

    def test_txt_as_xlsx(self):
        """Plain text file passed with .xlsx extension should raise exception."""
        with tempfile.NamedTemporaryFile(mode="w", suffix=".xlsx", delete=False) as f:
            f.write("This is plain text, not an xlsx file.\n")
            f.write("No ZIP magic bytes here.\n")
            tmp = f.name
        try:
            with pytest.raises(Exception):  # Could be BadZipFile or OSError
                self.processor.extract_xlsx(tmp)
        finally:
            os.unlink(tmp)

    def test_html_as_xlsx(self):
        """HTML file with .xlsx extension should raise exception."""
        with tempfile.NamedTemporaryFile(mode="w", suffix=".xlsx", delete=False) as f:
            f.write("<html><body><h1>Not an Excel file</h1></body></html>\n")
            tmp = f.name
        try:
            with pytest.raises(Exception):
                self.processor.extract_xlsx(tmp)
        finally:
            os.unlink(tmp)

    def test_binary_garbage_as_xlsx(self):
        """Random binary garbage as .xlsx should raise exception."""
        with tempfile.NamedTemporaryFile(mode="wb", suffix=".xlsx", delete=False) as f:
            f.write(b"\x00\x01\x02\x03\x04\x05" + b"NOT_A_VALID_XLSX_AT_ALL" * 100)
            tmp = f.name
        try:
            with pytest.raises(Exception):
                self.processor.extract_xlsx(tmp)
        finally:
            os.unlink(tmp)

    def test_truncated_zip_header_as_xlsx(self):
        """Truncated ZIP header (xlsx is a zip) should raise exception."""
        with tempfile.NamedTemporaryFile(mode="wb", suffix=".xlsx", delete=False) as f:
            # xlsx files are ZIPs; write only partial ZIP header
            f.write(b"PK\x03\x04")  # local file header signature
            f.write(b"\x14\x00" * 10)  # partial header fields
            tmp = f.name
        try:
            with pytest.raises(Exception):
                self.processor.extract_xlsx(tmp)
        finally:
            os.unlink(tmp)

    # ─── BOUNDARY: Valid xlsx structure, extreme data ───────────────────────

    def _create_xlsx_bytes(self, sheets_data: list) -> bytes:
        """Create a minimal xlsx file from sheets_data: list of (sheet_name, rows)."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheets_data[0][0] if sheets_data else "Sheet"
        for name, rows in sheets_data:
            if name != ws.title:
                ws = wb.create_sheet(name)
            for row in rows:
                ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_valid_xlsx_single_sheet(self):
        """Happy path: valid xlsx with single sheet returns non-empty string."""
        data = [("Data", [["A", "B", "C"], ["1", "2", "3"]])]
        xlsx_bytes = self._create_xlsx_bytes(data)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(xlsx_bytes)
            tmp = f.name
        try:
            result = self.processor.extract_xlsx(tmp)
            assert isinstance(result, str)
            assert len(result) > 0
            assert "Sheet: Data" in result
        finally:
            os.unlink(tmp)

    def test_valid_xlsx_multiple_sheets(self):
        """Happy path: valid xlsx with multiple sheets returns combined text."""
        data = [
            ("Sheet1", [["foo", "bar"]]),
            ("Sheet2", [["baz", "qux"]]),
            ("Sheet3", [["empty sheet placeholder"]]),
        ]
        xlsx_bytes = self._create_xlsx_bytes(data)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(xlsx_bytes)
            tmp = f.name
        try:
            result = self.processor.extract_xlsx(tmp)
            assert "Sheet: Sheet1" in result
            assert "Sheet: Sheet2" in result
            assert "Sheet: Sheet3" in result
        finally:
            os.unlink(tmp)

    def test_xlsx_empty_sheet_only(self):
        """xlsx with sheet containing only empty cells returns empty string."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, None, None])  # All empty row
        ws.append([])  # Completely empty row
        buf = io.BytesIO()
        wb.save(buf)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.getvalue())
            tmp = f.name
        try:
            result = self.processor.extract_xlsx(tmp)
            assert result == ""
        finally:
            os.unlink(tmp)

    def test_xlsx_many_rows(self):
        """xlsx with many rows (stress test) should not crash and returns non-empty."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(5000):
            ws.append([f"Row{i}", f"Data{i}", i])
        buf = io.BytesIO()
        wb.save(buf)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.getvalue())
            tmp = f.name
        try:
            result = self.processor.extract_xlsx(tmp)
            assert isinstance(result, str)
            assert len(result) > 0
        finally:
            os.unlink(tmp)

    def test_xlsx_many_sheets(self):
        """xlsx with 50 sheets should not crash."""
        import openpyxl
        wb = openpyxl.Workbook()
        for i in range(50):
            ws = wb.create_sheet(f"Sheet_{i:03d}")
            ws.append([f"Data in sheet {i}"])
        buf = io.BytesIO()
        wb.save(buf)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.getvalue())
            tmp = f.name
        try:
            result = self.processor.extract_xlsx(tmp)
            assert isinstance(result, str)
            assert "Sheet: Sheet_000" in result
            assert "Sheet: Sheet_049" in result
        finally:
            os.unlink(tmp)

    # ─── BOUNDARY: Unicode / Special Characters ──────────────────────────────

    def test_xlsx_unicode_sheet_names(self):
        """xlsx with unicode sheet names should not crash."""
        import openpyxl
        wb = openpyxl.Workbook()
        names = ["Привет", "こんにちは", "🎉", "العربية", "Emojisheet 🎊"]
        for name in names:
            ws = wb.create_sheet(name)
            ws.append(["data"])
        buf = io.BytesIO()
        wb.save(buf)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.getvalue())
            tmp = f.name
        try:
            result = self.processor.extract_xlsx(tmp)
            assert isinstance(result, str)
            # Should contain sheet headers for all sheets
            assert "Sheet: 🎉" in result
        finally:
            os.unlink(tmp)

    def test_xlsx_unicode_cell_content(self):
        """xlsx with unicode cell content should not crash."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Hello", "Привет", "こんにちは", "<script>alert(1)</script>", "null_byte_literal"])
        buf = io.BytesIO()
        wb.save(buf)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.getvalue())
            tmp = f.name
        try:
            result = self.processor.extract_xlsx(tmp)
            assert isinstance(result, str)
            # The HTML injection string should be preserved as-is (not executed)
            assert "<script>alert(1)</script>" in result
            assert "null_byte_literal" in result
        finally:
            os.unlink(tmp)

    # ─── BOUNDARY: Cell type edge cases ──────────────────────────────

    def test_xlsx_formula_cells(self):
        """xlsx formula cells (data_only=True) should return cached values, not crash."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=A1+A2"
        buf = io.BytesIO()
        wb.save(buf)
        # Re-open with data_only=True to get cached values
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.getvalue())
            tmp = f.name
        try:
            result = self.processor.extract_xlsx(tmp)
            assert isinstance(result, str)
            # A1 and A2 cached values should be present
            assert "10" in result
            assert "20" in result
        finally:
            os.unlink(tmp)

    def test_xlsx_date_cells(self):
        """xlsx with date cells should not crash."""
        import openpyxl
        from datetime import datetime
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([datetime(2024, 1, 15), "Label", 3.14159])
        buf = io.BytesIO()
        wb.save(buf)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.getvalue())
            tmp = f.name
        try:
            result = self.processor.extract_xlsx(tmp)
            assert isinstance(result, str)
            assert len(result) > 0
        finally:
            os.unlink(tmp)

    def test_xlsx_mixed_none_cells(self):
        """xlsx with None/empty cells interspersed with data should not crash."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, "Has Data", None, "More Data", None])
        ws.append(["A", None, "C", None, "E"])
        buf = io.BytesIO()
        wb.save(buf)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.getvalue())
            tmp = f.name
        try:
            result = self.processor.extract_xlsx(tmp)
            assert isinstance(result, str)
            assert "Has Data" in result
            assert "More Data" in result
        finally:
            os.unlink(tmp)

    # ─── BOUNDARY: Very large strings in cells ─────────────────────────────

    def test_xlsx_large_cell_content(self):
        """xlsx with very large cell content should not crash."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        large_text = "A" * 100_000
        ws.append([large_text])
        buf = io.BytesIO()
        wb.save(buf)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.getvalue())
            tmp = f.name
        try:
            result = self.processor.extract_xlsx(tmp)
            assert isinstance(result, str)
            # The large text should be present (truncated in str(cell))
            assert len(result) > 0
        finally:
            os.unlink(tmp)

    # ─── INTEGRATION: extract_document routes to extract_xlsx ──────────────

    def test_extract_document_routes_xlsx(self):
        """extract_document with .xlsx extension should call extract_xlsx."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["routed", "correctly"])
        buf = io.BytesIO()
        wb.save(buf)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.getvalue())
            tmp = f.name
        try:
            text, pages = self.processor.extract_document(tmp)
            assert isinstance(text, str)
            assert len(text) > 0
            assert pages == []  # xlsx returns empty pages list
        finally:
            os.unlink(tmp)

    def test_extract_document_xlsx_nonexistent_raises(self):
        """extract_document with non-existent .xlsx should raise FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            self.processor.extract_document("nonexistent_file.xlsx")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
